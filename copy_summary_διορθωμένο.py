
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

def copy_updated_sheets(template_path, target_path, output_path):
    # Load the template and target workbooks
    template_wb = load_workbook(template_path, data_only=False)
    target_wb = load_workbook(target_path)

    # Sheet mapping: source sheet -> destination name
    sheet_mappings = {
        'v1': 'Γενικό Αποτέλεσμα',
        'v2': 'Διαφορά'
    }

    for source, target in sheet_mappings.items():
        if target in target_wb.sheetnames:
            del target_wb[target]

        source_ws = template_wb[source]
        new_ws = target_wb.create_sheet(title=target)

        for row in source_ws.iter_rows():
            for cell in row:
                if isinstance(cell, Cell):
                    new_cell = new_ws.cell(row=cell.row, column=cell.column)
                    # Fix formula references from 'v1' to 'Γενικό Αποτέλεσμα'
                    if cell.data_type == "f":
                        formula = str(cell.value)
                        if source == "v2":
                            formula = formula.replace("'v1'", "'Γενικό Αποτέλεσμα'")
                        new_cell.value = f"={formula}" if not formula.startswith("=") else formula
                    else:
                        new_cell.value = cell.value
                    if cell.has_style:
                        new_cell._style = cell._style

    target_wb.save(output_path)

# Example usage:
# copy_updated_sheets('template.xlsx', 'client_file.xlsx', 'client_file_with_summaries.xlsx')
