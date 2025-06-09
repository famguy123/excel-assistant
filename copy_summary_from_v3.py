
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.styles import PatternFill

def copy_updated_sheets_with_formatting(template_path, target_path, output_path):
    template_wb = load_workbook(template_path, data_only=False)
    target_wb = load_workbook(target_path)

    sheet_mappings = {
        'v1': 'Γενικό Αποτέλεσμα',
        'v2': 'Διαφορά'
    }

    for source, target in sheet_mappings.items():
        if target in target_wb.sheetnames:
            del target_wb[target]

        source_ws = template_wb[source]
        new_ws = target_wb.create_sheet(title=target)

        for row in source_ws.iter_rows():
            for cell in row:
                if isinstance(cell, Cell):
                    new_cell = new_ws.cell(row=cell.row, column=cell.column)
                    if cell.data_type == "f":
                        formula = str(cell.value)
                        if source == "v2":
                            formula = formula.replace("'v1'", "'Γενικό Αποτέλεσμα'")
                        new_cell.value = f"={formula}" if not formula.startswith("=") else formula
                    else:
                        new_cell.value = cell.value
                    if cell.has_style:
                        new_cell._style = cell._style

    # Apply slightly darker blue fill
    formatting_ws = target_wb['Γενικό Αποτέλεσμα']
    cols = ["C", "E", "G", "I", "K", "M", "O", "Q", "S", "U", "W", "Y"]
    row_ranges = [
        range(36, 61), range(64, 89), range(96, 124), range(128, 158),
        range(162, 202), range(207, 236), range(240, 270), range(273, 313)
    ]
    fill = PatternFill(start_color="6699FF", end_color="6699FF", fill_type="solid")
    for rows in row_ranges:
        for r in rows:
            for col in cols:
                formatting_ws[f"{col}{r}"].fill = fill

    target_wb.save(output_path)
